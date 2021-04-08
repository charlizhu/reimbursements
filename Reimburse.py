from openpyxl import load_workbook
import shutil as sh
from datetime import date
from zipfile import ZipFile,ZIP_DEFLATED
import os
from tkinter import Tk
from tkinter.filedialog import askopenfilename

class Spreadsheets(object):

    def __init__(self):

        self.whichfolder = ""
        self.newname = r"eTransfer-Form-Fill.xlsx"
        self.reimbursements_form = load_workbook(filename=sh.copyfile(r"eTransfer-Form-Blank.xlsx",self.newname))
        self.eTransfer_sheet = self.reimbursements_form.active
        self.contacts_form = load_workbook(filename="Execs_List.xlsx")
        self.contacts_sheet = self.contacts_form.active
        self.today = date.today()
        self.name = ""
        self.student_id = ""
        self.email = ""
        self.phone = ""
        self.account_code = "591-7681-10"
        self.club_name = "UBC Mechanical Engineering Club"
        self.check_user()

    def check_user(self):
        user_name = input("Enter your name: ")
        for row in self.contacts_sheet.iter_rows(1):
            for recorded_name in row:
                if recorded_name.value == user_name:
                    self.name = str(row[0].value)
                    self.student_id = str(row[1].value)
                    self.phone = str(row[2].value)
                    self.email = str(row[3].value)
                else:
                    print("This is if the exec is not currently registered. Implement later.")
        self.eTransfer_sheet["B7"] = self.club_name
        self.eTransfer_sheet["B8"] = self.today
        self.eTransfer_sheet["B24"] = self.name
        self.eTransfer_sheet["B25"] = self.email
        self.eTransfer_sheet["B26"] = self.email
        self.eTransfer_sheet["B27"] = self.phone
        self.eTransfer_sheet["B28"] = self.student_id

        self.add_items()

    def add_items(self):
        flag = True
        count = 11
        current_directory = os.getcwd()
        final_directory = os.path.join(current_directory, r'Receipts')
        if not os.path.exists(final_directory):
            os.makedirs(final_directory)
        self.whichfolder = final_directory

        while flag:
            another = input("Do you have another reimbursement? (Y/N): ")
            if another == "N" or another == "n":
                flag = False
                break
            else:
                Tk().withdraw()
                sh.move(askopenfilename(), final_directory)
                description = input("In 10 words or less, enter a description for this item: ")
                amount = float(input("What is the CAD amount requested? "))
                self.eTransfer_sheet["A"+str(count)] = description
                self.eTransfer_sheet["B"+str(count)] = self.account_code
                self.eTransfer_sheet["C"+str(count)] = amount
                count+=1

        self.savefile()

    def savefile(self):
        self.reimbursements_form.save(filename=self.newname)
        zipObj = ZipFile(self.name + str(self.today) + '.zip', 'w')
        zipObj.write("eTransfer-Form-Fill.xlsx")
        for root, dirs, files in os.walk(self.whichfolder):
            for file in files:
                zipObj.write(os.path.join(root, file),
                           os.path.relpath(os.path.join(root, file),
                                           os.path.join(self.whichfolder, '..')))

        zipObj.close()


mysheet = Spreadsheets()
